from enum import Enum

from openpyxl.styles import PatternFill


class FilterFlag(Enum):
    # Marks a data point (e.g., contraction or overview) for deletion.
    # Styled with a red fill to indicate removal in Excel outputs.
    DELETE = PatternFill(start_color="ff7e75", end_color="FF0000", fill_type="solid")
    # Marks a data point for update (e.g., flagged as an outlier but not removed).
    # Styled with a light yellow fill to visually distinguish in Excel outputs.
    UPDATE = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
