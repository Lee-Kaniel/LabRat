import os
from typing import List, Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.worksheet.worksheet import Worksheet

from LabRat.src.configuration import OVERVIEW_FORMULAS
from LabRat.src.schemas.contraction import Contraction
from LabRat.src.schemas.filter_flag import FilterFlag
from LabRat.src.schemas.overview import Overview, sort_key
from LabRat.src.schemas.overview_table import OverviewTable
from LabRat.src.utils.excel import serialize_excel_worksheet_name, serialize_excel_title
from LabRat.src.utils.model_alias_functions import model_aliases, model_alias_count, get_fields_without_aliases


class Downloader:
    """
        Class responsible for downloading the processed data into Excel.
        Applies formatting, grouping, and filtering logic based on user configuration.
    """
    # Define header styles for Excel columns
    HEADER_STYLE = Font(color="FFFFFF", bold=True)
    HEADER_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

     # Define a built-in Excel table style for visual formatting
    TABLE_STYLE = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                 showRowStripes=True, showColumnStripes=True)
    
    # Combined headers from both Overview and Contraction schemas
    TABLE_HEADERS = model_aliases(Overview) + model_aliases(Contraction)

    def get_group_spreadsheet(self, wb: Workbook, group: str) -> Worksheet:
        """
        Returns a worksheet corresponding to a specific group name.
        If it doesn't exist, creates and formats it.
        """
        title = serialize_excel_worksheet_name(group)
        if title in wb.sheetnames:
            return wb[title]
        else:
            ws = wb.create_sheet(title=title)

            # If it's the first sheet delete the default sheet
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 2:
                del wb["Sheet"]

            # Add spreadsheet headers
            ws.append(self.TABLE_HEADERS)

            # Apply header cell style to each header cell
            for cell in ws[1]:
                cell.font = self.HEADER_STYLE
                cell.fill = self.HEADER_FILL

            return ws

    def format_finalized_workbook(self, wb: Workbook) -> None:
        """
        Adds Excel table formatting to each worksheet in the workbook.
        """
        for sheet_name in wb.sheetnames:
            ws: Worksheet = wb[sheet_name]
            table_ref = f"A1:{chr(ord('A') + ws.max_column - 1)}{ws.max_row + 1}"
            tab = Table(displayName=serialize_excel_title(ws.title), ref=table_ref)

            tab.tableStyleInfo = self.TABLE_STYLE
            ws.add_table(tab)

    @staticmethod
    def append_formula_row(
            ws: Worksheet, start_row: int, start_col: int, end_col: int, count: int, formula: str, index: int = 0
    ) -> None:
        """
        Appends a row with Excel formulas across the contraction columns for a given statistical operation.
        """
        for col in range(start_col, end_col + 1):
            start_cell = f"{chr(col + 64)}{start_row - count - index}"
            end_cell = f"{chr(col + 64)}{start_row - 1 - index}"
            cell_formula = f"={formula}({start_cell}:{end_cell})"

            ws.cell(row=start_row, column=col, value=cell_formula)

    def _write_overview_formulas_to_excel(
            self, overview_fields: List[Any], contraction_count: int, ws: Worksheet
    ) -> None:
        """
        Appends a summary row below a set of contractions, applying each formula in OVERVIEW_FORMULAS.
        """
        ws.append([None]) # spacing row

        overview_fields.pop() # remove last overview field (typically filter flag)
        current_row = ws.max_row + 1
        start_col = len(overview_fields) + 2 # Skip overview columns and empty column
        end_col = start_col + model_alias_count(Contraction) - 1

        formula_list = OVERVIEW_FORMULAS
        for index, formula in enumerate(formula_list):
            ws.append(overview_fields + [formula])
            self.append_formula_row(
                ws=ws, start_row=current_row + index, start_col=start_col, end_col=end_col,
                count=contraction_count + 1, formula=formula, index=index
            )

        ws.append([None])

    @staticmethod
    def _write_contraction_to_excel(
            contraction: Contraction, overview_fields: List[Any], ws: Worksheet, show_filtered: bool
    ) -> bool:
        """
            The function returns True if the Contraction was written to excel, False otherwise.

            If the show_filtered flag is set to False, all filtered Contractions will not be written to excel.
            If set to True all Contractions will be written, and filtered Contractions will be styled in red.
        """
        if show_filtered:
            contraction_values = list(contraction.dict(exclude=get_fields_without_aliases(Contraction)).values())
            ws.append([None for _ in overview_fields] + contraction_values)
            if filter_flag := contraction.filter_flag:
                row_index = len(ws["A"])
                for col_idx in range(len(overview_fields) + len(contraction_values)):
                    cell = ws.cell(row=row_index, column=col_idx + 1)
                    cell.fill = filter_flag.value
            return True
        else:
            if contraction.filter_flag != FilterFlag.DELETE:
                updated_contraction = contraction.model_copy(update=contraction.fields_for_update)
                contraction_values = list(updated_contraction.dict(
                    exclude=get_fields_without_aliases(Contraction)).values()
                )
                ws.append([None for _ in overview_fields] + contraction_values)
                return True
            return False

    def write_to_excel(
            self, data: OverviewTable, folder_path: str, show_filtered: bool = True, excel_file_name: str = None
    ) -> None:
        """
        Main method to write the full dataset to an Excel workbook.
        Iterates over grouped Overviews, applies filtering logic, and adds formula summaries.
        Saves final workbook to disk with appropriate naming conventions.
        """
        wb = Workbook()

        # Sort overviews for consistent ordering
        data.overview_list.sort(key=sort_key)
        for overview in data.overview_list:
            # Skip deleted overviews if exporting only unfiltered data
            if overview.filter_flag == FilterFlag.DELETE and not show_filtered:
                continue

            # Create or retrieve group worksheet
            ws = self.get_group_spreadsheet(wb=wb, group=overview.group)
            overview_fields = list(overview.dict(exclude=get_fields_without_aliases(Overview)).values())
            ws.append(overview_fields)

            # Style overview row if it has a filter flag
            if filter_flag := overview.filter_flag:
                row_index = len(ws["A"])
                for col_idx in range(len(overview_fields)):
                    ws.cell(row=row_index, column=len(overview_fields), value=overview.filter_flag.name)
                    cell = ws.cell(row=row_index, column=col_idx + 1)
                    cell.fill = filter_flag.value

            contraction_count = 0
            for contraction in overview.contraction_list:
                contraction_count += int(
                    self._write_contraction_to_excel(contraction, overview_fields, ws, show_filtered)
                )

            # Add summary formulas (e.g., averages) under the contraction data
            self._write_overview_formulas_to_excel(
                overview_fields=overview_fields, contraction_count=contraction_count, ws=ws
            )

        # Apply final Excel table formatting
        self.format_finalized_workbook(wb=wb)

        # Determine output file name
        if not excel_file_name:
            excel_file_name = f"{data.name}.xlsx" if show_filtered else f"{data.name}_filtered.xlsx"

        excel_file_path = os.path.join(folder_path, excel_file_name)
        wb.save(excel_file_path)
